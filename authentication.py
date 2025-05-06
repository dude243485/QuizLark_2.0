import os
import json
import re
from openpyxl import load_workbook
import random
import math

#regex password creation [A-Za-z0-9@#!%+=]
def checkPasswordStrength(s):
    upperCase = bool(re.search(r'[A-Z]', s))
    lowerCase = bool(re.search(r'[a-z]', s))
    digits = bool(re.search(r"\d", s))
    special = bool(re.search(r"[^A-Za-z0-9]", s))

    if upperCase and lowerCase and digits and special:
        return True

    return False

user_db_file = "user database.json"
#loading users from the database 
def load_users():
    """this function returns an object of the users database"""
    #return an empty dictionary if the path doesn't exist
    if not os.path.exists(user_db_file):
        return {}
    #open the file and return an object
    with open(user_db_file, "r") as file:
        return json.load(file)
    
def save_user(user):
    with open(user_db_file, "r") as file:
        users = json.load(file)
        users.append(user)
        
    with open(user_db_file, "w") as file: 
        json.dump(users, file, indent = 4)
        
def get_next_id():
    with open(user_db_file, "r") as file:
        users = json.load(file)
        return len(users)


def get_id(username):
    users = load_users()
    for user in users:
        if user["username"] == username:
            return user["id"]    


def register_user(username, password):
    users = load_users()
    if username in users:
        print("Username already exists. Choose another one")
        return False
    users[username] = password
    save_user(users)
    print("user registered sucessfully")
    return True


#check if user exists
def check_user(username):
    """this takes just one username parameter"""
    
    users = load_users()
    for user in users:
        if (user["username"] == username):
            return True
    
    return False

#check details      
def login_user(username, password):
    """This takes two parameters username and password"""
    users = load_users()
    for user in users:
        if (user["username"] == username) and (user["password"] == password):
            return True       
            
    else:
        return False
    
def get_xp(playerId):
    users = load_users()
    for user in users:
        if user["id"] == playerId:
            return user["xp"]
        
def get_money(playerId):
    users = load_users()
    for user in users:
        if user["id"] == playerId:
            return user["money"]
        
def get_save_me(playerId):
    users = load_users()
    for user in users:
        if user["id"] == playerId:
            return user["save me"]
        

def get_instant(playerId):
    users = load_users()
    for user in users:
        if user["id"] == playerId:
            return user["instant"]
        
def get_eliminate(playerId):
    users = load_users()
    for user in users:
        if user["id"] == playerId:
            return user["eliminate"]
        
def get_hint(playerId):
    users = load_users()
    for user in users:
        if user["id"] == playerId:
            return user["hint"]
        
def increase_value(playerId, value, increment):
    users = load_users()
    for user in users:
        if user["id"] == playerId:
            user[f"{value}"] = user[f"{value}"] + increment

    with open(user_db_file, "w") as file: 
        json.dump(users, file, indent = 4)

def set_money(playerId, price):
    users = load_users()
    for user in users:
        if user["id"] == playerId:
            user["money"] = user["money"] - price
    
    with open(user_db_file, "w") as file: 
        json.dump(users, file, indent = 4)

def get_stat(playerId, stat_name):
    users = load_users()
    for user in users:
        if user["id"] == playerId:
            return user["stats"][f"{stat_name}"]
        
    
        
def set_stat(playerId, stat_name, value):
    users = load_users()
    for user in users:
        if user["id"] == playerId:
            
            user["stats"][f"{stat_name}"] = value
        
    with open(user_db_file, "w") as file: 
        json.dump(users, file, indent = 4)
        
        
        
def set_powerUp(playerId, powerUp_name, value):
    users = load_users()
    for user in users:
        if user["id"] == playerId:
            user[powerUp_name] = value
    with open(user_db_file, "w") as file: 
        json.dump(users, file, indent = 4)
        
        
def get_leaderboards():
    """this function checks the database and returns a ranked(xp) list of players"""
    users = load_users()
    top_list = []
    top_ids  = [[0,0]]
    for user in users:
        top_ids.append([user["id"], user["xp"]])

    def quicksort(rank_list):
        if len(rank_list) <= 1:
            return rank_list
        
        #choosing the last element as the pivot
        pivot = rank_list[-1]
        left = [x for x in rank_list[:-1] if x[1] <= pivot[1]] #xp smaller than the pivot
        right = [x for x in rank_list[:-1] if x[1] > pivot[1]] #xp bigger than the pivot

        return quicksort(right) + [pivot] + quicksort(left)
    

    #return the top 5 xp-ranked players
    top_five = quicksort(top_ids)[:4]
    for player in top_five:
        id = player[0]
        for  user in users:
            if user["id"] == id:
                top_list.append([user["username"], user["xp"], user["money"], user["name"], user["id"]])

    return top_list

def get_top_categories(playerId):
    """this function checks the database and returns a ranked(xp) list of players"""
    users = load_users()
    for user in users:
        if user["id"] == playerId:
            categories = user["categories progress"]

    total_list = []
    for key in categories:
        total_list.append([key, categories[key]])

    def quicksort(rank_list):
        if len(rank_list) <= 1:
            return rank_list
        
        #choosing the last element as the pivot
        pivot = rank_list[-1]
        left = [x for x in rank_list[:-1] if x[1] <= pivot[1]] #xp smaller than the pivot
        right = [x for x in rank_list[:-1] if x[1] > pivot[1]] #xp bigger than the pivot

        return quicksort(right) + [pivot] + quicksort(left)
    
    top_categories = quicksort(total_list)[:4]
    for items in top_categories:
        if items[1] == 0:
            top_categories = list(top_categories)
            top_categories.remove(items)

    #if all categories are at 0
    if not top_categories:
        return ["no progress"], [1]
    
    titles = []
    values = []
    for i in top_categories:
        titles.append(i[0])
        values.append(i[1])
    return titles, values

#questions file
excel_file_path = "Database/Questions.xlsx"
def get_total_categories():
    workbook = load_workbook(excel_file_path)
    #get the number of sheets
    return workbook.sheetnames

def get_game_progress(playerId):
    """this returns the current player game progress as a list"""
    users = load_users()
    for user in users:
        if user["id"] == playerId:
            categories = user["categories progress"]

    progress_list = []
    for key in categories:
        progress_list.append(categories[key])

    return progress_list

def get_questions(category, level):
    start = (15 *(level - 1)) + 1
    end = 15 * level
    workbook = load_workbook(excel_file_path)

    sheet = workbook[category]
    row_list = []

    if level == 1:
        x = 1
    else: x = 0

    #iterate over rows from start_index to end_index
    for row in sheet.iter_rows(min_row= start + x, max_row = end, values_only= True):
        row_list.append(list(row))

    return row_list

def get_question(QList):
    question_pack = random.choice(QList)
    position = QList.index(question_pack)
    question_pack = QList.pop(position)
    question = question_pack[0]
    option_list = question_pack[1:5]
    random.shuffle(option_list)
    answer = question_pack[6]
    hint = question_pack[5]
    return question, option_list, answer, hint

def get_levels(sheetName):
    workbook = load_workbook(excel_file_path)
    sheet = workbook[sheetName]
    row_count = sheet.max_row
    levels = int(row_count/15)
    return levels

def get_progress(playerId, category):
    users = load_users()
    for user in users:
        if user["id"] == playerId:  
            level = math.ceil(user["categories progress"][category]/ 100 * get_levels(category))
            return level + 1
        
def get_actual_progress(playerId, category):
    users = load_users()
    for user in users:
        if user["id"] == playerId:
            return user["categories progress"][category]
        
def reset_progress(playerId, category):
    users = load_users()
    for user in users:
        if user["id"] == playerId:
            user["categories progress"][category] = 0
            
    with open(user_db_file, "w") as file: 
        json.dump(users, file, indent = 4)   
            
    
        
    

def set_progress(playerId,category, level):
    users = load_users()
    for user in users:
        if user["id"] == playerId:
            user["categories progress"][category ] = round(((level/get_levels(category)) * 100), 2)
    with open(user_db_file, "w") as file: 
        json.dump(users, file, indent = 4)
        
    



    



    

    




    



 

   
            
        
    
    